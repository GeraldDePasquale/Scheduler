import csv
import time
from datetime import datetime, date, timedelta, time
from tkinter import Tk, filedialog, simpledialog
from openpyxl import workbook, worksheet, load_workbook

root = Tk()
directory = "C:\\ProgramData\\MathnasiumScheduler"
FILEOPENOPTIONS = dict(defaultextension='.csv', filetypes=[('XLSX', '*.xlsx'), ('CSV file', '*.csv')])

# define days consistent with datetime.weekday() for printing and sorting purposes
mon = 0
tue = 1
wed = 2
thu = 3
fri = 4
sat = 5
sun = 6

logfile = None
instructors = []
config_file_name = None
config_wb = None
config_ws = None
availability_file_name = None
availability_wb = None
availability_ws = None
instructor_schedule_file = None
instructor_ws = None

initialized = False

unavailable = time(0, 0)

# center instruction hours
instructionHours = {sun: [], \
                    mon: [time(15, 30), time(19, 30)], \
                    tue: [time(15, 30), time(19, 30)], \
                    wed: [time(15, 30), time(19, 30)], \
                    thu: [time(15, 30), time(19, 30)], \
                    fri: [], \
                    sat: [time(10, 00), time(14, 00)]}


class Instructor_from_Row:
    def __init__(self, row, prev=None, next=None):

        self.prev = prev
        self.next = next

        # define work schedule
        self.schedule = {sun: [], mon: [], tue: [], wed: [], thu: [], fri: [], sat: []}

        # set instructor name
        self.availability_reported = row[0].value
        self.email_address = row[1].value
        self.name = row[2].value

        # set instructors hours of availability based upon the lines in the InstructorAvailability input file
        self.availability = {sun: (time(0, 0), time(0, 0)), \
                             mon: (row[5].value, row[6].value), \
                             tue: (row[8].value, row[9].value), \
                             wed: (row[11].value, row[12].value), \
                             thu: (row[14].value, row[15].value), \
                             fri: (time(0, 0), time(0, 0)), \
                             sat: (row[17].value, row[18].value)}

        # set rank
        self.rank = 999 #Not found
        #find row of the instructor, then assign rank from cell
        first_row = 2  # skip the headers
        last_row = Instructor_from_Row.config_ws.max_row
        last_col = Instructor_from_Row.config_ws.max_column
        for row in Instructor_from_Row.config_ws.iter_rows(min_row=first_row, max_col=last_col, max_row=last_row):
            if row[0].value == self.name:
                self.rank = row[2].value

        print(Instructor_from_Row.config_ws["A1"].value)

        # set cost
        self.cost = 15

        # set max hours per week
        self.maxHrsPerWeek = 20

        # set min hours per day
        self.minHrsPerDay = 2

        # set work day preferences
        self.workDayPreferences = {sun: 0, mon: 2, tue: 3, wed: 4, thu: 5, fri: 6, sat: 7}
        # set dayofweek strings
        self.dayStrings = {sun: "Sunday", mon: "Monday", tue: "Tuesday", \
                           wed: "Wednesday", thu: "Thursday", fri: "Friday", \
                           sat: "Saturday"}

        # set isSchedule flag - this variable is set by the scheduling process
        self.isScheduled = False

    @classmethod
    def initialize(self, file):
        Instructor_from_Row.log_file = file
        # Open Configuration File
        Instructor_from_Row.config_file_name = filedialog.askopenfilename(parent=root, initialdir=directory,
                                                        title='Select Configuration File', **FILEOPENOPTIONS)
        Instructor_from_Row.config_wb = load_workbook(Instructor_from_Row.config_file_name,
                                                      data_only=True, guess_types=True)
        Instructor_from_Row.config_ws = Instructor_from_Row.config_wb.active
        print("Loaded ", Instructor_from_Row.config_file_name)

        # Open Instructor Availabilty File
        Instructor_from_Row.availability_file_name = filedialog.askopenfilename(parent=root, initialdir=directory,
                                                                       title='Select Instructor Availability File',
                                                                       **FILEOPENOPTIONS)
        Instructor_from_Row.availability_wb = load_workbook(Instructor_from_Row.availability_file_name,
                                                            data_only=True, guess_types=True)
        Instructor_from_Row.availability_ws = Instructor_from_Row.availability_wb.active
        print("Loaded ", Instructor_from_Row.availability_file_name)

        #Clean up Awailability Work Sheet
        # Replace periods of unavailability with valid time (eg.g None becomes time(0,0))
        first_row = 2  # skip the headers
        last_row = Instructor_from_Row.availability_ws.max_row
        last_col = Instructor_from_Row.availability_ws.max_column
        for row in Instructor_from_Row.availability_ws.iter_rows(min_row=first_row, max_col=last_col, max_row=last_row):
            for i in [5,6,8,9,11,12,14,15,17,18]:
                if row[i].value == None: row[i].value = time(0,0)
        for eachInstructor in instructors: print(eachInstructor.name)

        # Create Instructors
        Instructor_from_Row.instructors = []
        for row in Instructor_from_Row.availability_ws.iter_rows(min_row=first_row, max_col=last_col, max_row=last_row):
            Instructor_from_Row.instructors.append(Instructor_from_Row(row, Instructor_from_Row.log_file))
        for eachInstructor in Instructor_from_Row.instructors: print("Instructor: ", eachInstructor.name,
                                                                     "Rank: ", eachInstructor.rank)
        print("Created Instructors")

    # define instructor sort
    def __lt__(self, other):
        return self.rank < other.rank

    def __str__(self):
        return str(self.name)

    def __print__(self):
        print(str(self.name, self.rank))

    def prev(self):
        return self.prev

    def next(self):
        return self.next

    # weekdayString(self, dayofweek) returns string representation of the day
    def dayString(self, integer):
        return self.dayStrings[integer]

    # isAvailable returns true if the instructor is available to work at the time of the event
    def isAvailable(self, event):
        dayNeeded = event.eventTime.weekday()
        timeNeeded = event.eventTime.time()
        imScheduled = self.isScheduled
        availableStartTime = self.availability[dayNeeded][0]
        availableStopTime = self.availability[dayNeeded][1]
        result = (not imScheduled) and \
                 (timeNeeded >= availableStartTime) and \
                 (timeNeeded < availableStopTime)
        return result

    # isAvailableToOpen returns true if the instructor is available to start work when the center opens
    def isAvailableToOpen(self, event):
        dayNeeded = event.eventTime.weekday()
        timeNeeded = instructionHours[dayNeeded][0]
        imScheduled = self.isScheduled
        myStartTime = self.availability[dayNeeded][0]
        myStopTime = self.availability[dayNeeded][1]
        result = (not imScheduled) and \
                 (timeNeeded >= myStartTime) and \
                 (timeNeeded < myStopTime)
        return result

    # startWork adds the start work event time to the instructor's work schedule
    def startWork(self, event):
        dayNeeded = event.eventTime.weekday()
        timeNeeded = event.eventTime.time()
        self.schedule[dayNeeded].append(timeNeeded)
        self.isScheduled = True

    # startWorkWhenOpen adds start at open time to the instructor's work schedule
    def startWorkWhenOpen(self, event):
        dayNeeded = event.eventTime.weekday()
        timeNeeded = instructionHours[dayNeeded][0]
        self.schedule[dayNeeded].append(timeNeeded)
        self.isScheduled = True

    # stopWork adds the stop work event time to the instructor's work schedule
    def stopWork(self, event):
        self.schedule[event.eventTime.weekday()].append(event.eventTime.time())
        self.isScheduled = False

    # mustDepart returns true if the event time exceeds the instructors availability
    def mustDepart(self, event):
        dayNeeded = event.eventTime.weekday()
        timeNeeded = event.eventTime.time()
        myStopTime = self.availability[dayNeeded][1]
        result = (timeNeeded >= myStopTime)
        return result

    # departWork adds stop work event but leaves instructor in isScheduled state
    def departWork(self, event):
        self.schedule[event.eventTime.weekday()].append(event.eventTime.time())
        self.isScheduled = True

    # hoursScheduled returns the number of hours the employee is scheduled to work
    #    def hoursScheduledNow(self,event):

    # adjust time to land on the hour or half hour
    def adjustTime(self, aTime):
        if aTime.minute < 30:
            return time(aTime.hour, 0)
        elif aTime.minute > 30:
            return time(aTime.hour, 30)
        return aTime

    # finalizeSchedule modifies the schedule so that each work day includes one start and one stop time
    def finalizeSchedule(self):
        for eachKey in self.schedule.keys():
            if self.schedule[eachKey]:
                myStartTime = max(self.schedule[eachKey][0], self.availability[eachKey][0])
                myStopTime = min(self.availability[eachKey][1], self.schedule[eachKey][len(self.schedule[eachKey]) - 1])
                # Bug fix: If no stop work ordered, stop work at end of day. Better fix check for no stop work
                if myStopTime == myStartTime: myStopTime = instructionHours[eachKey][1]
                self.schedule[eachKey] = [self.adjustTime(myStartTime), self.adjustTime(myStopTime)]
        return self.schedule
