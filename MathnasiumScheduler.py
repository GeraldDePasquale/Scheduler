# Mathnasium Staffing Forecast
#   Forecasts instructor staffing requirements for scheduling purposes.
#   Forecasts based upon a student arrival-departure event model.

# Outputs:
# Excel Workbook containing:
# 1. Instructor Schedules
# 2. Summary Staffing Forecast Sheet provides number of instructors required each instruction day. Shows only
#   events when actual increases or decreases in staffing levels are required.
# 3. Detailed Stafford Forecast Sheet provides same data as Summary Staffing Forecast except that it provides
#    every event used for forcasting.

# Inputs:
# 1. Radius Attendance Report
# 2. Radius Student Report
# 3. Configuration File - includes instructor ranking for scheduling purposes. Also provides other instructor
# specific data (e.g. cost / hour)
# 4. Work Availability Spreadsheet - provides instructor availability as input using Google Form: Work Availability

# Hardcoded Parameters:
#    lowCost = 1/5 instructor to student ratio (Grades 2 .. 5)
#    mediumCost = 1/4 instructor to student ratio (Grades 6 .. 8)
#    highCost = 1/3 instructor to student ratio (Grades 0, 1, 9+)
#    veryHighCost = 1.0 private tutoring
#    churnTolerance = 360
#       The number of seconds per hour that will be tolerated to minimize
#       churn and avoid overstaffing due to students arrivals and departures.
#       For instance, if we need desire to maintain a 1:3 instructor:student
#       ratio and we have 10:31, how long will we tolerate it before
#       we bring in another instructor. Due to student churn, that ratio may
#       return to 10:30 in a matter of minutes. How many minutes are tolerable?
#       Default: 6 minutes per hour.

import math
import os
import os.path
from datetime import datetime
from tkinter import Tk, simpledialog
from openpyxl import Workbook
from Event import Event
from Instructor import Instructor
from Student import Student
from Importer import Importer
from Reporter import Reporter
from Gmailer import Gmailer
import httplib2
from googleapiclient.discovery import build
from oauth2client.client import AccessTokenCredentials
from GoogleCredentials import GoogleCredentials

def main():

    debug = False
    root = Tk()
    root.withdraw()
    run_time = datetime.now().strftime('%Y%m%d%H%M') # used for file extensions, makes sorting easy
    print("\nMathnasium Scheduler Launched")
    default_directory = "C:\\ProgramData\\MathnasiumScheduler"
    FILEOPENOPTIONS = dict(defaultextension='.csv', filetypes=[('XLSX', '*.xlsx'), ('CSV file', '*.csv')])
    # Todo-jerry add center name picklist, get center names from configuraton file
    center_name = simpledialog.askstring("Name prompt", "Enter Center Name")
    schedule_start_date = simpledialog.askstring("Schedule Start Prompt", "Enter Schedule Start Date (MM/DD/YY)")
    # center_name = "aaaa.TestRun" #Eliminates need to select files for successive test runs
    importer = Importer().import_all(run_time, default_directory, center_name, FILEOPENOPTIONS)

    #Create Schedule Workbook
    # Create Run Workbook
    run_wb_path = default_directory + "\\" + center_name + "." + run_time + ".xlsx"
    run_wb = Workbook()
    run_wb.save(run_wb_path)
    schedule_by_name_ws = run_wb.create_sheet("Schedule By Name", index=0)
    schedule_by_day_ws = run_wb.create_sheet("Schedule By Day", index=1)
    forecast_summary_ws = run_wb.create_sheet("Summary Forecast", index=2)
    forecast_detailed_ws = run_wb.create_sheet("Detailed Forecast", index=3)
    run_log_ws = run_wb.create_sheet("Runlog", index=4)
    #ToDo Write run_log to run_log_ws
    run_log = []
    instructors = Instructor.create_instructors(run_log)
    students = Student.initialize_students(importer.attendance_ws, importer.student_data_ws, run_log)

    #Create Events
    print("\nCreating events from student arrivals and departures\n")
    events = []
    for each_student in Student.students:
        events.append(Event('Arrival', each_student.arrival_time, each_student))
        events.append(Event('Departure', each_student.departure_time, each_student))
    events.sort()

    #Executing Events
    print("Executing events and collecting information")
    # Gather events by week day
    #ToDo refactor this into Common and use common
    # define days consistent with datetime.weekday()
    mon = 0
    tue = 1
    wed = 2
    thu = 3
    fri = 4
    sat = 5
    sun = 6

    # Group events by day
    event_groups = {sun: [], mon: [], tue: [], wed: [], thu: [], fri: [], sat: []}
    for each_event in events:
        event_groups[each_event.event_time.weekday()].append(each_event)

    print("\tDetermining cost of each day")
    costsOfEventGroups = {}
    for each_day in event_groups.keys():
        cost = 0.0
        for each_event in event_groups[each_day]:
            if each_event.is_arrival_event: cost = cost + each_event.cost()
        costsOfEventGroups[each_day] = round(cost, 1)
        print("\t\tDay: ", each_day, "Cost: ", costsOfEventGroups[each_day])

    # Sort and process each group of events
    for each_day in event_groups.keys():
        print("\n\tProcessing Day: ", str(each_day))
        event_groups[each_day].sort()
        instructorsMinimum = 2.0  # ToDo remove hard coded variable the minimum staffing level
        instructorsRequired = 0.0  # actual number of instructors required to meet student demand
        event_number = 1  # first event number
        student_count = 0  # start with zero students
        for each_event in event_groups[each_day]:
            # Set event number
            each_event.event_number = event_number
            # Set event's previous and next events
            if event_number != 1: each_event.prev = events[events.index(each_event) - 1]
            if event_number != len(event_groups[each_day]): each_event.next = event_groups[each_day][
                event_groups[each_day].index(each_event) + 1]
            event_number = event_number + 1  # next event number
            # Maintain student count
            if (each_event.is_arrival_event):
                student_count = student_count + 1
            elif (each_event.is_departure_event):
                student_count = student_count - 1
            each_event.student_count = student_count
            # Compute/maintain the actual number of instructors required
            instructorsRequired = instructorsRequired + each_event.cost()
            # Compute/maintain the number of instructors to staff (minimum is instructorsMinimum)
            each_event.instructor_count = max(instructorsMinimum, math.ceil(instructorsRequired))

        print("\t\tCollecting Instructor Change Events")
        instructor_change_events = []
        for each_event in event_groups[each_day]:
            if each_event.is_instructor_change_event():
                instructor_change_events.append(each_event)

        print("\t\tMarking Churn Events")
        tolerance = 360  # seconds (6 minutes)
        for i in range(len(instructor_change_events) - 1):
            event = instructor_change_events[i]
            next_event = instructor_change_events[i + 1]
            event.isChurnEvent = event.is_peak_event() and next_event.is_valley_event() \
                                 and (next_event.event_time - event.event_time).seconds < tolerance

        print("\t\tScheduling Instructors")
#        instructors = Instructor.instructors
        instructors.sort()
        if debug:
            for eachInstructor in instructors: print(eachInstructor)
        inactive_instructors = instructors
        active_instructors = []
        dateChangeEvents = 0

        for each_event in event_groups[each_day]:
            if each_event.is_date_change_event():
                dateChangeEvents = dateChangeEvents + 1
                # Activate minimum number of instructors needed to open if necessary
                active_instructor_count = 0
                inactive_instructors.sort()
                instructors_changed = []
                for this_instructor in inactive_instructors:
                    if this_instructor.isAvailableToOpen(each_event) and (active_instructor_count < instructorsMinimum):
                        active_instructor_count = active_instructor_count + 1
                        this_instructor.startWorkWhenOpen(each_event)
                        active_instructors.append(this_instructor)
                        # save pointers to instructors for removal from unscheduled list
                        instructors_changed.append(this_instructor)
                for i in instructors_changed: inactive_instructors.remove(i)

            # Check for and remove departing Instructors
            departed_instructors = []
            for this_instructor in active_instructors:
                departed = False
                if this_instructor.mustDepart(each_event):
                    this_instructor.departWork(each_event)
                    departed_instructors.append(this_instructor)

            instructor_change_needed = each_event.instructor_count - len(active_instructors)
            if not each_event.is_churn_event and instructor_change_needed > 0:
                # Schedule available instructors
                # print("\t\t\tActivate Instructor")
                active_instructor_count = 0
                inactive_instructors.sort()
                # print("Inactive Instructors: " + str(len(inactive_instructors)))
                while (active_instructor_count < instructor_change_needed):
                    # print('while')
                    # Find instructor and schedule instructor
                    instructors_changed = [] #ToDo inside the while loop (inconsistent see line 215)
                    for this_instructor in inactive_instructors:
                        if this_instructor.isAvailable(each_event) and (active_instructor_count < instructor_change_needed):
                            active_instructor_count = active_instructor_count + 1
                            this_instructor.startWork(each_event)
                            active_instructors.append(this_instructor)
                            # Save pointers to newly scheduled instructors for removal from unscheduled list
                            instructors_changed.append(this_instructor)
                # Remove newly activated (scheduled) instructors from inactive list
                for i in instructors_changed: inactive_instructors.remove(i)

            if not each_event.is_churn_event and instructor_change_needed < 0:
                # Deactivate instructors
                # print("\t\t\tDeactivate Instructor")
                deactivated_instructor_count = 0
                instructors_deactivated = [] #ToDo outside the while loop (inconsistent see line 200)
                active_instructors.sort()
                # print("Printing Reversed Rank List")
                # for this in reversed(active_instructors):
                #     print(this.name, this.rank)
                while deactivated_instructor_count < abs(instructor_change_needed):
                    for this_instructor in reversed(active_instructors):
                          if deactivated_instructor_count < abs(instructor_change_needed):
                            deactivated_instructor_count = deactivated_instructor_count + 1
                            this_instructor.stopWork(each_event)
                            inactive_instructors.append(this_instructor)
                            instructors_deactivated.append(this_instructor)
                for i in instructors_deactivated: active_instructors.remove(i)

            # Finalize schedules after last departure or final event of the day
            if (each_event == event_groups[each_day][len(event_groups[each_day]) - 1]) or each_event.next.is_date_change_event():
                instructors_changed = []
                for this_instructor in active_instructors:
                    this_instructor.isScheduled = False
                    inactive_instructors.append(this_instructor)
                    instructors_changed.append(this_instructor)
                for i in instructors_changed: active_instructors.remove(i)
                for i in inactive_instructors: i.finalizeSchedule()

    Reporter().write_all(events, instructors, forecast_detailed_ws, forecast_summary_ws, schedule_by_name_ws)

    print("\nReview, edit, and approve schedules")
    # Todo code to review, edit, and approve schedules

    print("\nFormating and Closing Workbooks")
    Importer().close_workbooks()
    Reporter().format_sheets(run_wb)
    run_wb.save(run_wb_path)
    run_wb.close()

    print("\nLaunching Excel")
    os.system("start excel " + run_wb_path )

    # Todo code up individual schedule emails including mapping to email addresses and instructor first names.
    print("\nEmailing Schedules to Instructors")
#    Gmailer().send_instructor_schedules(instructors)

    # Todo code up scheduling individual work events on master schedule calendar and instructor calendars.
    print("\nAdding Work Events to Instructor Google Calendars")
#    GoogleEventScheduler().insert_events(instructors)

    print("\nScheduler Run Completed")

if __name__ == '__main__': main()
