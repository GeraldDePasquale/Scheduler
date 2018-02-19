from tkinter import Tk, filedialog
from openpyxl import load_workbook

root = Tk()
root.withdraw()


class Importer:

    instructor_availability_wb = None
    instructor_availability_ws = None
    student_data_wb = None
    student_data_ws = None
    attendance_wb = None
    attendance_ws = None
    config_wb = None
    config_ws = None

    # Todo-jerry remove hard coded names and directories
    root_directory = None
    working_copy = "Working Copy"

    @classmethod
    def import_all(cls, run_time, directory, center_name, FILEOPENOPTIONS):
        # test file names
        student_attendance_file_name = "C:\\ProgramData\\MathnasiumScheduler\\Stafford 20180211/Stafford Attendance Radius.20180211.xlsx"
        student_data_file_name = "C:\\ProgramData/MathnasiumScheduler\\Stafford 20180211\\Stafford.Student Report.20180214.xlsx"
        config_file_name = "C:\\ProgramData\\MathnasiumScheduler/Stafford 20180211\\Stafford Configuration File.20180211.xlsx"
        instructor_availability_file_name = "C:\\ProgramData\\MathnasiumScheduler\\Stafford 20180211/Stafford Instructor Availability.20180211.xlsx"

        importer = cls()
        # Load Attendance Report
        # Source: Radius>>Reports>>Student Attendance Type: .xlsx
#        student_attendance_file_name = filedialog.askopenfilename(parent=root, initialdir=directory,
#                                                                  title='Select Radius Student Attendance Report',
#                                                                  **FILEOPENOPTIONS)
        Importer.attendance_wb = load_workbook(student_attendance_file_name, data_only=True, guess_types=True)
        Importer.attendance_ws = Importer.attendance_wb.active
        print("Imported Attendance Report", student_attendance_file_name)

        # Load Student Report
        # Source: Radius>>Reports>>Student  Type: .xlsx
#        student_data_file_name = filedialog.askopenfilename(parent=root, initialdir=directory,
#                                                            title='Select Student Report',
#                                                            **FILEOPENOPTIONS)
        Importer.student_data_wb = load_workbook(student_data_file_name, data_only=True, guess_types=True)
        Importer.student_data_ws = Importer.student_data_wb.active
        print("Imported Student Report", student_data_file_name)

        # Load Configuration
#        config_file_name = filedialog.askopenfilename(parent=root, initialdir=directory,
#                                                      title='Select Configuration File',
#                                                      **FILEOPENOPTIONS)
        Importer.config_wb = load_workbook(config_file_name, data_only=True, guess_types=True)
        Importer.config_ws = Importer.config_wb.active

        print("Loaded ", config_file_name)

        # Load Instructor Availabilty
#        instructor_availability_file_name = filedialog.askopenfilename(parent=root, initialdir=directory,
#                                                                       title='Select Instructor Availability File',
#                                                                       **FILEOPENOPTIONS)
        Importer.instructor_availability_wb = load_workbook(instructor_availability_file_name, data_only=True, guess_types=True)
        Importer.instructor_availability_ws = Importer.instructor_availability_wb.active
        print("Loaded ", instructor_availability_file_name)
        return importer

    @staticmethod
    def close_workbooks():
        Importer.attendance_wb.close()
        Importer.student_data_wb.close()
        Importer.config_wb.close()
        Importer.instructor_availability_wb.close()
